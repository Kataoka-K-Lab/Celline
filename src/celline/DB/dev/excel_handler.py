"""
Excel-based metadata storage handler with validation.
Manages metadata in metadata.xlsx with projects, samples, and runs sheets.
"""

from __future__ import annotations
from typing import Dict, List, Optional, Any, Set
from dataclasses import dataclass, field
import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ValidationError:
    """Represents a validation error."""
    sheet: str
    row: Optional[int]
    column: Optional[str]
    error_type: str
    message: str


@dataclass
class ValidationResult:
    """Result of validation with errors and warnings."""
    is_valid: bool
    errors: List[ValidationError] = field(default_factory=list)
    warnings: List[ValidationError] = field(default_factory=list)

    def add_error(self, sheet: str, row: Optional[int], column: Optional[str],
                  error_type: str, message: str) -> None:
        """Add a validation error."""
        self.errors.append(ValidationError(sheet, row, column, error_type, message))
        self.is_valid = False

    def add_warning(self, sheet: str, row: Optional[int], column: Optional[str],
                    error_type: str, message: str) -> None:
        """Add a validation warning."""
        self.warnings.append(ValidationError(sheet, row, column, error_type, message))


class ExcelMetadataHandler:
    """Handler for Excel-based metadata storage with validation."""

    # Required columns for each sheet
    REQUIRED_COLUMNS = {
        "available_samples": ["sample_id", "added_date"],
        "projects": ["key", "title", "summary", "children", "parent"],
        "samples": ["key", "title", "summary", "species", "raw_link", "parent", "children"],
        "runs": ["key", "title", "strategy", "raw_link", "parent", "children"],
    }

    # Required sheets (available_samples must be first)
    REQUIRED_SHEETS = ["available_samples", "projects", "samples", "runs"]

    def __init__(self, excel_path: str):
        """
        Initialize Excel metadata handler.

        Args:
            excel_path: Path to metadata.xlsx file
        """
        self.excel_path = excel_path
        self._ensure_file_exists()

    def _ensure_file_exists(self) -> None:
        """Ensure Excel file exists, create if not."""
        if not os.path.exists(self.excel_path):
            self._create_empty_excel()

    def _create_empty_excel(self) -> None:
        """Create a new Excel file with required sheets and columns."""
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create required sheets with headers
        for sheet_name in self.REQUIRED_SHEETS:
            ws = wb.create_sheet(sheet_name)
            columns = self.REQUIRED_COLUMNS[sheet_name]
            for idx, col in enumerate(columns, 1):
                ws.cell(row=1, column=idx, value=col)

        wb.save(self.excel_path)

    def validate_structure(self) -> ValidationResult:
        """
        Validate the Excel file structure (sheets and columns).

        Returns:
            ValidationResult with any structural errors
        """
        result = ValidationResult(is_valid=True)

        if not os.path.exists(self.excel_path):
            result.add_error("file", None, None, "FileNotFound",
                           f"Excel file not found: {self.excel_path}")
            return result

        try:
            wb = load_workbook(self.excel_path)
        except Exception as e:
            result.add_error("file", None, None, "InvalidFile",
                           f"Cannot open Excel file: {str(e)}")
            return result

        # Check required sheets exist
        existing_sheets = set(wb.sheetnames)
        for required_sheet in self.REQUIRED_SHEETS:
            if required_sheet not in existing_sheets:
                result.add_error("file", None, None, "MissingSheet",
                               f"Required sheet '{required_sheet}' not found")

        # Check columns for each sheet
        for sheet_name in self.REQUIRED_SHEETS:
            if sheet_name not in existing_sheets:
                continue

            ws = wb[sheet_name]
            if ws.max_row < 1:
                result.add_error(sheet_name, None, None, "EmptySheet",
                               "Sheet has no header row")
                continue

            # Get header row
            headers = [cell.value for cell in ws[1]]
            headers_set = set(h.lower() if h else "" for h in headers)

            # Check required columns
            required_cols = set(col.lower() for col in self.REQUIRED_COLUMNS[sheet_name])
            missing_cols = required_cols - headers_set

            if missing_cols:
                result.add_error(sheet_name, 1, None, "MissingColumns",
                               f"Missing required columns: {', '.join(missing_cols)}")

        return result

    def validate_data_integrity(self) -> ValidationResult:
        """
        Validate data integrity including:
        - Non-empty keys (primary keys)
        - Unique keys within each sheet
        - Parent-child relationship consistency
        - Referenced IDs exist

        Returns:
            ValidationResult with any integrity errors
        """
        result = ValidationResult(is_valid=True)

        # First validate structure
        structure_result = self.validate_structure()
        if not structure_result.is_valid:
            result.errors.extend(structure_result.errors)
            return result

        # Load all sheets
        try:
            projects_df = pd.read_excel(self.excel_path, sheet_name="projects")
            samples_df = pd.read_excel(self.excel_path, sheet_name="samples")
            runs_df = pd.read_excel(self.excel_path, sheet_name="runs")
        except Exception as e:
            result.add_error("file", None, None, "ReadError",
                           f"Cannot read Excel file: {str(e)}")
            return result

        # Normalize column names to lowercase
        projects_df.columns = [c.lower() if isinstance(c, str) else c for c in projects_df.columns]
        samples_df.columns = [c.lower() if isinstance(c, str) else c for c in samples_df.columns]
        runs_df.columns = [c.lower() if isinstance(c, str) else c for c in runs_df.columns]

        # 1. Check for empty keys
        for sheet_name, df in [("projects", projects_df), ("samples", samples_df), ("runs", runs_df)]:
            if "key" not in df.columns:
                continue

            for idx, row in df.iterrows():
                if pd.isna(row["key"]) or str(row["key"]).strip() == "":
                    result.add_error(sheet_name, idx + 2, "key", "EmptyKey",
                                   "Key cannot be empty")

        # 2. Check for duplicate keys
        for sheet_name, df in [("projects", projects_df), ("samples", samples_df), ("runs", runs_df)]:
            if "key" not in df.columns:
                continue

            keys = df["key"].dropna().astype(str)
            duplicates = keys[keys.duplicated()].unique()
            if len(duplicates) > 0:
                result.add_error(sheet_name, None, "key", "DuplicateKeys",
                               f"Duplicate keys found: {', '.join(duplicates)}")

        # Load available_samples for lenient validation
        try:
            available_samples_df = pd.read_excel(self.excel_path, sheet_name="available_samples")
            available_samples_df.columns = [c.lower() if isinstance(c, str) else c for c in available_samples_df.columns]
            available_sample_ids = set(available_samples_df["sample_id"].dropna().astype(str))
        except Exception:
            available_sample_ids = set()

        # Collect all valid keys
        project_keys = set(projects_df["key"].dropna().astype(str))
        sample_keys = set(samples_df["key"].dropna().astype(str))
        run_keys = set(runs_df["key"].dropna().astype(str))

        # 3. Validate parent-child relationships

        # Projects: children should be in available_samples (not necessarily in samples sheet)
        for idx, row in projects_df.iterrows():
            if pd.notna(row.get("children")):
                children_str = str(row["children"]).strip()
                if children_str:
                    child_ids = [c.strip() for c in children_str.split(",")]
                    for child_id in child_ids:
                        # Only error if not in available_samples (allows partial addition)
                        if child_id and child_id not in available_sample_ids:
                            result.add_warning("projects", idx + 2, "children", "UnavailableSample",
                                           f"Child sample '{child_id}' not in available_samples (not added yet)")

        # Samples: parent should be valid project ID, children should be valid run IDs
        for idx, row in samples_df.iterrows():
            # Check parent
            if pd.notna(row.get("parent")):
                parent_id = str(row["parent"]).strip()
                if parent_id and parent_id not in project_keys:
                    result.add_error("samples", idx + 2, "parent", "InvalidParent",
                                   f"Parent project '{parent_id}' not found in projects sheet")

            # Check children
            if pd.notna(row.get("children")):
                children_str = str(row["children"]).strip()
                if children_str:
                    child_ids = [c.strip() for c in children_str.split(",")]
                    for child_id in child_ids:
                        if child_id and child_id not in run_keys:
                            result.add_error("samples", idx + 2, "children", "InvalidChild",
                                           f"Child run '{child_id}' not found in runs sheet")

        # Runs: parent should be valid sample ID
        for idx, row in runs_df.iterrows():
            if pd.notna(row.get("parent")):
                parent_id = str(row["parent"]).strip()
                if parent_id and parent_id not in sample_keys:
                    result.add_error("runs", idx + 2, "parent", "InvalidParent",
                                   f"Parent sample '{parent_id}' not found in samples sheet")

        # 4. Check bidirectional consistency (warnings)

        # Check if samples listed in project.children have correct parent
        for idx, row in projects_df.iterrows():
            project_key = str(row["key"])
            if pd.notna(row.get("children")):
                children_str = str(row["children"]).strip()
                if children_str:
                    child_ids = [c.strip() for c in children_str.split(",")]
                    for child_id in child_ids:
                        if child_id in sample_keys:
                            sample_row = samples_df[samples_df["key"].astype(str) == child_id]
                            if not sample_row.empty:
                                parent = sample_row.iloc[0].get("parent")
                                if pd.isna(parent) or str(parent).strip() != project_key:
                                    result.add_warning("projects", idx + 2, "children", "InconsistentRelation",
                                                     f"Sample '{child_id}' parent does not match this project")

        # Check if runs listed in sample.children have correct parent
        for idx, row in samples_df.iterrows():
            sample_key = str(row["key"])
            if pd.notna(row.get("children")):
                children_str = str(row["children"]).strip()
                if children_str:
                    child_ids = [c.strip() for c in children_str.split(",")]
                    for child_id in child_ids:
                        if child_id in run_keys:
                            run_row = runs_df[runs_df["key"].astype(str) == child_id]
                            if not run_row.empty:
                                parent = run_row.iloc[0].get("parent")
                                if pd.isna(parent) or str(parent).strip() != sample_key:
                                    result.add_warning("samples", idx + 2, "children", "InconsistentRelation",
                                                     f"Run '{child_id}' parent does not match this sample")

        return result

    def read_sheet(self, sheet_name: str) -> pd.DataFrame:
        """
        Read a specific sheet from the Excel file.

        Args:
            sheet_name: Name of the sheet to read

        Returns:
            DataFrame with the sheet data
        """
        if sheet_name not in self.REQUIRED_SHEETS:
            raise ValueError(f"Invalid sheet name: {sheet_name}")

        if not os.path.exists(self.excel_path):
            return pd.DataFrame(columns=self.REQUIRED_COLUMNS[sheet_name])

        df = pd.read_excel(self.excel_path, sheet_name=sheet_name)
        df.columns = [c.lower() if isinstance(c, str) else c for c in df.columns]
        return df

    def write_sheet(self, sheet_name: str, df: pd.DataFrame) -> None:
        """
        Write data to a specific sheet.

        Args:
            sheet_name: Name of the sheet to write
            df: DataFrame with the data to write
        """
        if sheet_name not in self.REQUIRED_SHEETS:
            raise ValueError(f"Invalid sheet name: {sheet_name}")

        # Load existing workbook or create new
        if os.path.exists(self.excel_path):
            with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                for sheet in self.REQUIRED_SHEETS:
                    if sheet == sheet_name:
                        df.to_excel(writer, sheet_name=sheet, index=False)
                    else:
                        pd.DataFrame(columns=self.REQUIRED_COLUMNS[sheet]).to_excel(
                            writer, sheet_name=sheet, index=False)

    def get_record(self, sheet_name: str, key: str) -> Optional[Dict[str, Any]]:
        """
        Get a single record by key.

        Args:
            sheet_name: Name of the sheet
            key: Key to search for

        Returns:
            Dictionary with record data or None if not found
        """
        df = self.read_sheet(sheet_name)
        if "key" not in df.columns:
            return None

        matching = df[df["key"].astype(str) == str(key)]
        if matching.empty:
            return None

        return matching.iloc[0].to_dict()

    def add_or_update_record(self, sheet_name: str, record: Dict[str, Any]) -> None:
        """
        Add or update a record in the specified sheet.

        Args:
            sheet_name: Name of the sheet
            record: Dictionary with record data (must include 'key')
        """
        if "key" not in record:
            raise ValueError("Record must have a 'key' field")

        df = self.read_sheet(sheet_name)
        key = str(record["key"])

        # Remove existing record with same key
        if "key" in df.columns:
            df = df[df["key"].astype(str) != key]

        # Add new record
        new_df = pd.DataFrame([record])
        df = pd.concat([df, new_df], ignore_index=True)

        self.write_sheet(sheet_name, df)

    def get_all_keys(self, sheet_name: str) -> Set[str]:
        """
        Get all keys from a sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            Set of all keys
        """
        df = self.read_sheet(sheet_name)
        if "key" not in df.columns or df.empty:
            return set()
        return set(df["key"].dropna().astype(str))

    def add_to_available_samples(self, sample_id: str) -> None:
        """
        Add a sample ID to available_samples sheet.

        Args:
            sample_id: Sample ID to add
        """
        from datetime import datetime

        df = self.read_sheet("available_samples")

        # Check if already exists
        if "sample_id" in df.columns:
            if sample_id in df["sample_id"].astype(str).values:
                return  # Already exists

        # Add new entry
        new_entry = pd.DataFrame([{
            "sample_id": sample_id,
            "added_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }])

        df = pd.concat([df, new_entry], ignore_index=True)
        self.write_sheet("available_samples", df)

    def get_available_samples(self) -> Set[str]:
        """
        Get all available sample IDs.

        Returns:
            Set of available sample IDs
        """
        df = self.read_sheet("available_samples")
        if "sample_id" not in df.columns or df.empty:
            return set()
        return set(df["sample_id"].dropna().astype(str))

    def add_multiple_to_available_samples(self, sample_ids: List[str]) -> None:
        """
        Add multiple sample IDs to available_samples sheet.

        Args:
            sample_ids: List of sample IDs to add
        """
        from datetime import datetime

        df = self.read_sheet("available_samples")

        # Get existing sample IDs
        existing_ids = set()
        if "sample_id" in df.columns and not df.empty:
            existing_ids = set(df["sample_id"].astype(str).values)

        # Filter out duplicates
        new_ids = [sid for sid in sample_ids if sid not in existing_ids]

        if not new_ids:
            return  # Nothing to add

        # Create new entries
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_entries = pd.DataFrame([
            {"sample_id": sid, "added_date": timestamp}
            for sid in new_ids
        ])

        df = pd.concat([df, new_entries], ignore_index=True)
        self.write_sheet("available_samples", df)
